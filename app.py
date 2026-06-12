import streamlit as st
import anthropic
import base64
import json
import pandas as pd
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import io

st.set_page_config(
    page_title="Rider Performance Tracker",
    page_icon="🛵",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

*, html, body, [class*="css"] {
  font-family: 'Inter', sans-serif !important;
  box-sizing: border-box;
}

/* ── dark background ── */
.stApp { background: #080b14 !important; }
section[data-testid="stSidebar"] { background: #0d1020 !important; }
.block-container { padding: 0 !important; max-width: 100% !important; }

/* ── hide streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none; }

/* ── top navbar ── */
.navbar {
  background: linear-gradient(90deg, #0d1020 0%, #111827 100%);
  border-bottom: 1px solid #1e2540;
  padding: 14px 32px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  position: sticky;
  top: 0;
  z-index: 999;
}
.navbar-brand { display: flex; align-items: center; gap: 12px; }
.navbar-brand span { font-size: 22px; font-weight: 800; color: #fff; letter-spacing: -.5px; }
.navbar-brand .dot { color: #6366f1; }
.navbar-badge {
  background: #1e2540;
  border: 1px solid #2a3050;
  border-radius: 20px;
  padding: 4px 14px;
  font-size: 12px;
  color: #8b92b0;
  font-weight: 500;
}
.navbar-right { display: flex; gap: 10px; align-items: center; }

/* ── main content wrapper ── */
.dash-wrap { padding: 24px 32px; }

/* ── section title ── */
.sec-title {
  font-size: 13px;
  font-weight: 600;
  color: #6366f1;
  text-transform: uppercase;
  letter-spacing: .1em;
  margin: 28px 0 14px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.sec-title::after {
  content: '';
  flex: 1;
  height: 1px;
  background: linear-gradient(90deg, #1e2540, transparent);
}

/* ── metric cards ── */
.metrics-grid {
  display: grid;
  grid-template-columns: repeat(5, 1fr);
  gap: 14px;
  margin-bottom: 8px;
}
.metric-card {
  background: linear-gradient(135deg, #111827 0%, #0d1020 100%);
  border: 1px solid #1e2540;
  border-radius: 14px;
  padding: 20px 18px;
  position: relative;
  overflow: hidden;
  transition: border-color .2s, transform .15s;
}
.metric-card:hover { border-color: #6366f1; transform: translateY(-2px); }
.metric-card::before {
  content: '';
  position: absolute;
  top: 0; left: 0; right: 0;
  height: 3px;
  background: var(--accent, #6366f1);
  border-radius: 14px 14px 0 0;
}
.mc-icon { font-size: 22px; margin-bottom: 10px; }
.mc-label { font-size: 11px; font-weight: 600; color: #6b7280; text-transform: uppercase; letter-spacing: .08em; }
.mc-value { font-size: 32px; font-weight: 800; color: #fff; margin: 4px 0 2px; line-height: 1; }
.mc-sub { font-size: 12px; color: var(--accent, #6366f1); font-weight: 500; }

/* ── upload panel ── */
.upload-panel {
  background: #0d1020;
  border: 1px solid #1e2540;
  border-radius: 16px;
  padding: 24px;
  margin-bottom: 20px;
}
.paste-zone {
  background: #111827;
  border: 2px dashed #2a3050;
  border-radius: 12px;
  padding: 32px 24px;
  text-align: center;
  cursor: pointer;
  transition: border-color .2s, background .2s;
  margin-bottom: 16px;
  position: relative;
}
.paste-zone:hover { border-color: #6366f1; background: #131929; }
.paste-zone-icon { font-size: 36px; margin-bottom: 10px; }
.paste-zone-text { color: #6b7280; font-size: 14px; }
.paste-zone-text strong { color: #a5b4fc; }
.paste-preview {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(120px, 1fr));
  gap: 10px;
  margin-top: 16px;
}
.paste-thumb {
  border-radius: 8px;
  border: 1px solid #1e2540;
  overflow: hidden;
  aspect-ratio: 16/9;
  background: #111827;
}
.paste-thumb img { width: 100%; height: 100%; object-fit: cover; }

div[data-testid="stFileUploader"] {
  background: transparent !important;
  border: none !important;
  padding: 0 !important;
}
div[data-testid="stFileUploader"] > div {
  background: #111827 !important;
  border: 2px dashed #2a3050 !important;
  border-radius: 12px !important;
  padding: 24px !important;
}
div[data-testid="stFileUploader"] label { color: #6b7280 !important; }

/* ── buttons ── */
div.stButton > button {
  background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  padding: 12px 28px !important;
  font-weight: 600 !important;
  font-size: 14px !important;
  width: 100% !important;
  letter-spacing: .02em !important;
  transition: opacity .2s, transform .15s !important;
  box-shadow: 0 4px 15px rgba(99,102,241,.3) !important;
}
div.stButton > button:hover {
  opacity: .9 !important;
  transform: translateY(-1px) !important;
}

/* ── API key input ── */
div[data-testid="stTextInput"] input {
  background: #111827 !important;
  border: 1px solid #1e2540 !important;
  border-radius: 10px !important;
  color: #fff !important;
  padding: 10px 14px !important;
  font-size: 14px !important;
}
div[data-testid="stTextInput"] input:focus {
  border-color: #6366f1 !important;
  box-shadow: 0 0 0 3px rgba(99,102,241,.15) !important;
}
div[data-testid="stTextInput"] label { color: #6b7280 !important; font-size: 12px !important; font-weight: 600 !important; text-transform: uppercase !important; letter-spacing: .08em !important; }

/* ── text area (paste) ── */
div[data-testid="stTextArea"] textarea {
  background: #111827 !important;
  border: 2px dashed #2a3050 !important;
  border-radius: 12px !important;
  color: #fff !important;
  font-size: 14px !important;
  min-height: 120px !important;
}

/* ── rider table ── */
.rider-table {
  background: #0d1020;
  border: 1px solid #1e2540;
  border-radius: 16px;
  overflow: hidden;
}
.rt-header {
  display: grid;
  grid-template-columns: 200px 80px 120px 80px 100px 100px 110px 110px 120px;
  background: #111827;
  border-bottom: 1px solid #1e2540;
  padding: 12px 20px;
}
.rt-hcell {
  font-size: 11px;
  font-weight: 600;
  color: #6b7280;
  text-transform: uppercase;
  letter-spacing: .08em;
}
.rt-row {
  display: grid;
  grid-template-columns: 200px 80px 120px 80px 100px 100px 110px 110px 120px;
  padding: 14px 20px;
  border-bottom: 1px solid #0f1526;
  align-items: center;
  transition: background .15s;
}
.rt-row:hover { background: #111827; }
.rt-row:last-child { border-bottom: none; }
.rt-cell { font-size: 13px; color: #cbd5e1; }
.rt-name { font-weight: 600; color: #fff; font-size: 14px; }
.rt-id { font-size: 11px; color: #6b7280; margin-top: 2px; }

/* ── badges ── */
.badge {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  border-radius: 20px;
  padding: 3px 10px;
  font-size: 11px;
  font-weight: 600;
}
.badge-working { background: rgba(74,222,128,.1); color: #4ade80; border: 1px solid rgba(74,222,128,.2); }
.badge-offline { background: rgba(248,113,113,.1); color: #f87171; border: 1px solid rgba(248,113,113,.2); }
.badge-idle    { background: rgba(251,191,36,.1);  color: #fbbf24; border: 1px solid rgba(251,191,36,.2); }

/* ── value colors ── */
.val-green  { color: #4ade80; font-weight: 700; }
.val-yellow { color: #fbbf24; font-weight: 700; }
.val-blue   { color: #60a5fa; font-weight: 700; }
.val-white  { color: #fff;    font-weight: 700; }
.val-purple { color: #a78bfa; font-weight: 700; }

/* ── alerts ── */
.alert {
  border-radius: 10px;
  padding: 12px 16px;
  font-size: 13px;
  margin: 12px 0;
  display: flex;
  align-items: center;
  gap: 10px;
}
.alert-success { background: rgba(74,222,128,.08); border: 1px solid rgba(74,222,128,.2); color: #4ade80; }
.alert-error   { background: rgba(248,113,113,.08); border: 1px solid rgba(248,113,113,.2); color: #f87171; }
.alert-info    { background: rgba(99,102,241,.08);  border: 1px solid rgba(99,102,241,.2);  color: #a5b4fc; }

/* ── progress ── */
div[data-testid="stProgress"] > div { background: #1e2540 !important; border-radius: 10px !important; }
div[data-testid="stProgress"] > div > div { background: linear-gradient(90deg, #4f46e5, #7c3aed) !important; border-radius: 10px !important; }

/* ── download button ── */
div[data-testid="stDownloadButton"] button {
  background: linear-gradient(135deg, #059669 0%, #0d9488 100%) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  font-weight: 600 !important;
  box-shadow: 0 4px 15px rgba(5,150,105,.25) !important;
}

/* ── dataframe ── */
.stDataFrame { border-radius: 12px !important; overflow: hidden !important; }
iframe[title="st.iframe"] { border-radius: 12px; }

/* ── tabs ── */
.stTabs [data-baseweb="tab-list"] {
  background: #0d1020 !important;
  border-radius: 10px !important;
  gap: 4px !important;
  padding: 4px !important;
  border: 1px solid #1e2540 !important;
}
.stTabs [data-baseweb="tab"] {
  background: transparent !important;
  color: #6b7280 !important;
  border-radius: 8px !important;
  font-weight: 600 !important;
  font-size: 13px !important;
}
.stTabs [aria-selected="true"] {
  background: #1e2540 !important;
  color: #fff !important;
}
.stTabs [data-baseweb="tab-panel"] { background: transparent !important; padding-top: 20px !important; }

/* ── empty state ── */
.empty-state {
  text-align: center;
  padding: 80px 20px;
  background: #0d1020;
  border: 1px dashed #1e2540;
  border-radius: 16px;
}
.empty-icon { font-size: 56px; margin-bottom: 16px; }
.empty-title { font-size: 20px; font-weight: 700; color: #fff; margin-bottom: 8px; }
.empty-sub { font-size: 14px; color: #6b7280; max-width: 360px; margin: 0 auto; }

/* scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0d1020; }
::-webkit-scrollbar-thumb { background: #2a3050; border-radius: 3px; }
</style>
""", unsafe_allow_html=True)

# ── session state ─────────────────────────────────────────────────────────────
for k, v in [("all_riders", []), ("paste_images", []), ("last_updated", None)]:
    if k not in st.session_state:
        st.session_state[k] = v

# ── helpers ───────────────────────────────────────────────────────────────────

def img_to_b64(img_bytes: bytes) -> tuple[str, str]:
    """Returns (base64_str, media_type)."""
    # detect format
    try:
        img = Image.open(io.BytesIO(img_bytes))
        fmt = (img.format or "PNG").upper()
        mt_map = {"JPEG": "image/jpeg", "JPG": "image/jpeg",
                  "PNG": "image/png", "WEBP": "image/webp", "GIF": "image/gif"}
        mt = mt_map.get(fmt, "image/png")
    except Exception:
        mt = "image/png"
    return base64.standard_b64encode(img_bytes).decode(), mt


def extract_riders(client, img_b64: str, media_type: str) -> list[dict]:
    prompt = """You are a data extraction assistant for a delivery operations dashboard.

Analyze this screenshot carefully and extract ALL riders/couriers visible.

For EACH rider return a JSON object with EXACTLY these keys:
- rider_id: numeric ID as string
- name: full display name
- session: session time e.g. "18:00 - 00:00"
- utr: UTR float e.g. 0.6
- deliveries: system deliveries count integer (the number next to "Deliveries")
- accepted: REAL accepted orders integer (from donut chart "Accepted" label - this is the TRUE count)
- stacked: stacked orders integer (0 if not shown)
- acceptance_rate: float e.g. 50.0
- cash_balance: float numeric only e.g. 197.36
- currency: "SAR", "EGP", "USD" etc.
- rider_state: "Working" | "Offline" | "Idle"
- location_area: area text or ""
- phone: phone number string or ""

IMPORTANT: "deliveries" is what the SYSTEM shows (often wrong/low).
"accepted" is the REAL count from the donut/pie chart accepted section.

Return ONLY a raw JSON array, no markdown, no code fences, no explanation.
If no riders: return []"""

    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
            {"type": "text", "text": prompt}
        ]}]
    )
    raw = resp.content[0].text.strip().replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def merge_riders(existing: list, new: list) -> list:
    id_map = {r["rider_id"]: i for i, r in enumerate(existing) if r.get("rider_id")}
    result = existing.copy()
    for r in new:
        rid = r.get("rider_id", "")
        if rid and rid in id_map:
            result[id_map[rid]] = r
        else:
            result.append(r)
            if rid:
                id_map[rid] = len(result) - 1
    return result


def build_excel(riders: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riders Performance"

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color="4F46E5")
    c = Alignment(horizontal="center", vertical="center")
    l = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="E5E7EB")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ID", "Name", "Phone", "Session", "UTR",
               "Sys. Deliveries", "Real Accepted", "Stacked",
               "Acceptance %", "Cash Balance", "Currency", "State", "Location"]
    widths   = [12, 28, 16, 16, 8, 16, 14, 10, 14, 14, 10, 12, 22]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = c; cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    wfill = PatternFill("solid", start_color="ECFDF5")
    ofill = PatternFill("solid", start_color="FEF2F2")
    afill = PatternFill("solid", start_color="FEFCE8")
    a2    = PatternFill("solid", start_color="F9FAFB")

    for ri, r in enumerate(riders, 2):
        state = str(r.get("rider_state", "")).lower()
        rfill = wfill if "work" in state else ofill if "off" in state else afill if "idle" in state else (a2 if ri % 2 == 0 else None)
        row_data = [
            r.get("rider_id",""), r.get("name",""), r.get("phone",""),
            r.get("session",""), r.get("utr",""),
            r.get("deliveries",""), r.get("accepted",""), r.get("stacked",""),
            r.get("acceptance_rate",""), r.get("cash_balance",""),
            r.get("currency",""), r.get("rider_state",""), r.get("location_area","")
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bdr
            cell.alignment = l if ci == 2 else c
            if rfill: cell.fill = rfill
        ws.row_dimensions[ri].height = 22

    # summary sheet
    ws2 = wb.create_sheet("Summary")
    utrs_v = [r["utr"] for r in riders if isinstance(r.get("utr"), (int,float))]
    summary = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Riders", len(riders)),
        ("Working", sum(1 for r in riders if "work" in str(r.get("rider_state","")).lower())),
        ("Offline",  sum(1 for r in riders if "off"  in str(r.get("rider_state","")).lower())),
        ("Avg UTR",  round(sum(utrs_v)/len(utrs_v), 2) if utrs_v else 0),
        ("Total Real Accepted", sum(r.get("accepted",0) or 0 for r in riders)),
        ("Total Sys. Deliveries", sum(r.get("deliveries",0) or 0 for r in riders)),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def badge_html(state: str) -> str:
    s = state.lower()
    if "work" in s: return '<span class="badge badge-working">● Working</span>'
    if "off"  in s: return '<span class="badge badge-offline">● Offline</span>'
    return '<span class="badge badge-idle">● Idle</span>'


# ── get API key from secrets or input ─────────────────────────────────────────
api_key = st.secrets.get("ANTHROPIC_API_KEY", "") if hasattr(st, "secrets") else ""

# ── NAVBAR ────────────────────────────────────────────────────────────────────
total_r   = len(st.session_state.all_riders)
working_r = sum(1 for r in st.session_state.all_riders if "work" in str(r.get("rider_state","")).lower())

st.markdown(f"""
<div class="navbar">
  <div class="navbar-brand">
    <span>🛵 Rider<span class="dot">.</span>Tracker</span>
    <span class="navbar-badge">{total_r} riders · {working_r} working</span>
  </div>
  <div class="navbar-right">
    <span class="navbar-badge">v2.0</span>
  </div>
</div>
<div class="dash-wrap">
""", unsafe_allow_html=True)

# ── TABS ──────────────────────────────────────────────────────────────────────
tab_upload, tab_dashboard, tab_table = st.tabs(["📤  Upload & Extract", "📊  Dashboard", "📋  Full Table"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
with tab_upload:

    # API key (only if not in secrets)
    if not api_key:
        st.markdown('<div class="sec-title">🔑 API Configuration</div>', unsafe_allow_html=True)
        api_key = st.text_input("Anthropic API Key", type="password", placeholder="sk-ant-api03-...")
        st.markdown('<div class="alert alert-info">💡 To skip entering the key every time, add <b>ANTHROPIC_API_KEY</b> in your Streamlit Secrets (Settings → Secrets)</div>', unsafe_allow_html=True)

    # ── upload section ────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">📎 Add Screenshots</div>', unsafe_allow_html=True)

    input_tab1, input_tab2 = st.tabs(["📁  Upload Files", "📋  Paste Image"])

    collected_images = []  # list of (bytes, filename)

    with input_tab1:
        uploaded_files = st.file_uploader(
            "Drop screenshots here",
            type=["png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            label_visibility="collapsed"
        )
        if uploaded_files:
            for f in uploaded_files:
                f.seek(0)
                collected_images.append((f.read(), f.name))
            st.markdown(f'<div class="alert alert-info">📎 {len(uploaded_files)} file(s) ready</div>', unsafe_allow_html=True)

    with input_tab2:
        st.markdown("""
        <div style="color:#6b7280; font-size:13px; margin-bottom:12px;">
        📋 Copy a screenshot to clipboard (Ctrl+C / Cmd+C) then paste it below using Ctrl+V / Cmd+V in the text field — or paste a URL of the image.
        </div>""", unsafe_allow_html=True)

        paste_url = st.text_input(
            "Paste image URL or base64",
            placeholder="https://... or data:image/png;base64,...",
            label_visibility="collapsed"
        )
        if paste_url:
            import urllib.request
            try:
                if paste_url.startswith("data:image"):
                    # base64 data URI
                    header, b64data = paste_url.split(",", 1)
                    img_bytes = base64.b64decode(b64data)
                    collected_images.append((img_bytes, "pasted_image.png"))
                    st.markdown('<div class="alert alert-success">✅ Image loaded from base64</div>', unsafe_allow_html=True)
                elif paste_url.startswith("http"):
                    req = urllib.request.Request(paste_url, headers={"User-Agent": "Mozilla/5.0"})
                    with urllib.request.urlopen(req, timeout=10) as resp:
                        img_bytes = resp.read()
                    collected_images.append((img_bytes, "url_image.png"))
                    st.markdown('<div class="alert alert-success">✅ Image loaded from URL</div>', unsafe_allow_html=True)
            except Exception as e:
                st.markdown(f'<div class="alert alert-error">❌ Could not load image: {e}</div>', unsafe_allow_html=True)

        # clipboard paste via file uploader fallback
        st.markdown('<div style="color:#6b7280;font-size:12px;margin-top:12px;">Or paste directly — click below then Ctrl+V:</div>', unsafe_allow_html=True)
        pasted_file = st.file_uploader(
            "Paste here",
            type=["png","jpg","jpeg","webp"],
            key="paste_uploader",
            label_visibility="collapsed"
        )
        if pasted_file:
            pasted_file.seek(0)
            collected_images.append((pasted_file.read(), "pasted_" + pasted_file.name))
            st.markdown('<div class="alert alert-success">✅ Pasted image ready</div>', unsafe_allow_html=True)

    # preview thumbnails
    if collected_images:
        st.markdown('<div class="sec-title">🖼️ Preview</div>', unsafe_allow_html=True)
        cols = st.columns(min(len(collected_images), 6))
        for i, (img_bytes, fname) in enumerate(collected_images):
            with cols[i % 6]:
                try:
                    st.image(img_bytes, caption=fname[:20], use_container_width=True)
                except Exception:
                    st.write(fname)

    # ── action buttons ────────────────────────────────────────────────────────
    st.markdown("")
    c1, c2, c3 = st.columns([3, 1, 1])
    with c1:
        extract_btn = st.button("🔍  Extract Rider Data", use_container_width=True)
    with c2:
        clear_btn = st.button("🗑️  Clear All", use_container_width=True)
    with c3:
        if st.session_state.all_riders:
            xlsx = build_excel(st.session_state.all_riders)
            fname = f"riders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button("📥  Export", data=xlsx, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    if clear_btn:
        st.session_state.all_riders = []
        st.rerun()

    # ── extraction logic ──────────────────────────────────────────────────────
    if extract_btn:
        if not api_key:
            st.markdown('<div class="alert alert-error">❌ Enter your Anthropic API key above.</div>', unsafe_allow_html=True)
        elif not collected_images:
            st.markdown('<div class="alert alert-error">❌ Add at least one screenshot first.</div>', unsafe_allow_html=True)
        else:
            client = anthropic.Anthropic(api_key=api_key)
            prog = st.progress(0, text="Initializing…")
            new_riders = []
            errors = []

            for i, (img_bytes, fname) in enumerate(collected_images):
                prog.progress(i / len(collected_images), text=f"Reading {fname}…")
                try:
                    b64, mt = img_to_b64(img_bytes)
                    riders = extract_riders(client, b64, mt)
                    new_riders.extend(riders)
                except Exception as e:
                    errors.append(f"{fname}: {e}")

            prog.progress(1.0, text="Done ✓")
            st.session_state.all_riders = merge_riders(st.session_state.all_riders, new_riders)
            st.session_state.last_updated = datetime.now().strftime("%H:%M:%S")

            if new_riders:
                st.markdown(f'<div class="alert alert-success">✅ Extracted <b>{len(new_riders)}</b> rider(s) from <b>{len(collected_images)}</b> image(s) — Total in report: <b>{len(st.session_state.all_riders)}</b></div>', unsafe_allow_html=True)
            for err in errors:
                st.markdown(f'<div class="alert alert-error">⚠️ {err}</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
with tab_dashboard:
    riders = st.session_state.all_riders

    if not riders:
        st.markdown("""
        <div class="empty-state">
          <div class="empty-icon">🛵</div>
          <div class="empty-title">No data yet</div>
          <div class="empty-sub">Go to the Upload tab, add your screenshots, and click Extract.</div>
        </div>""", unsafe_allow_html=True)
    else:
        # ── KPI cards ─────────────────────────────────────────────────────────
        working_c  = sum(1 for r in riders if "work" in str(r.get("rider_state","")).lower())
        offline_c  = sum(1 for r in riders if "off"  in str(r.get("rider_state","")).lower())
        utrs_v     = [r["utr"] for r in riders if isinstance(r.get("utr"),(int,float))]
        avg_utr    = round(sum(utrs_v)/len(utrs_v), 2) if utrs_v else 0
        tot_acc    = sum(r.get("accepted",0) or 0 for r in riders)
        tot_del    = sum(r.get("deliveries",0) or 0 for r in riders)
        diff       = tot_acc - tot_del

        st.markdown(f"""
        <div class="metrics-grid">
          <div class="metric-card" style="--accent:#6366f1">
            <div class="mc-icon">👥</div>
            <div class="mc-label">Total Riders</div>
            <div class="mc-value">{len(riders)}</div>
            <div class="mc-sub">in current report</div>
          </div>
          <div class="metric-card" style="--accent:#4ade80">
            <div class="mc-icon">🟢</div>
            <div class="mc-label">Working</div>
            <div class="mc-value">{working_c}</div>
            <div class="mc-sub">active right now</div>
          </div>
          <div class="metric-card" style="--accent:#f87171">
            <div class="mc-icon">🔴</div>
            <div class="mc-label">Offline</div>
            <div class="mc-value">{offline_c}</div>
            <div class="mc-sub">not on shift</div>
          </div>
          <div class="metric-card" style="--accent:#60a5fa">
            <div class="mc-icon">⚡</div>
            <div class="mc-label">Avg UTR</div>
            <div class="mc-value">{avg_utr}</div>
            <div class="mc-sub">utilization rate</div>
          </div>
          <div class="metric-card" style="--accent:#a78bfa">
            <div class="mc-icon">✅</div>
            <div class="mc-label">Real Accepted</div>
            <div class="mc-value">{tot_acc}</div>
            <div class="mc-sub">+{diff} vs system</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── last updated ──────────────────────────────────────────────────────
        if st.session_state.last_updated:
            st.markdown(f'<div style="color:#6b7280;font-size:12px;text-align:right;margin-bottom:8px;">Last updated: {st.session_state.last_updated}</div>', unsafe_allow_html=True)

        # ── rider cards ───────────────────────────────────────────────────────
        st.markdown('<div class="sec-title">🏍️ Riders</div>', unsafe_allow_html=True)

        # sort by UTR desc
        sorted_riders = sorted(riders, key=lambda r: r.get("utr", 0) or 0, reverse=True)

        for r in sorted_riders:
            bdg  = badge_html(r.get("rider_state",""))
            curr = r.get("currency","")
            utr  = r.get("utr","-")
            acc  = r.get("accepted","-")
            dl   = r.get("deliveries","-")
            ar   = r.get("acceptance_rate","-")
            bal  = r.get("cash_balance",0)

            utr_color = "#4ade80" if (isinstance(utr,(int,float)) and utr >= 0.8) else \
                        "#fbbf24" if (isinstance(utr,(int,float)) and utr >= 0.5) else "#f87171"

            st.markdown(f"""
            <div class="rider-card" style="background:#0d1020;border:1px solid #1e2540;border-radius:14px;
                padding:16px 22px;margin-bottom:10px;display:flex;align-items:center;
                justify-content:space-between;transition:border-color .2s;"
                onmouseover="this.style.borderColor='#6366f1'"
                onmouseout="this.style.borderColor='#1e2540'">
              <div style="min-width:200px;">
                <div style="color:#fff;font-size:15px;font-weight:700;">{r.get('name','Unknown')}</div>
                <div style="color:#6b7280;font-size:12px;margin-top:3px;">
                  ID: {r.get('rider_id','-')} &nbsp;·&nbsp; 📞 {r.get('phone','—')}
                </div>
                <div style="color:#6b7280;font-size:11px;margin-top:2px;">
                  🕐 {r.get('session','—')}
                </div>
              </div>
              <div style="display:flex;gap:28px;align-items:center;flex-wrap:wrap;">
                <div style="text-align:center;">
                  <div style="color:#6b7280;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">UTR</div>
                  <div style="color:{utr_color};font-size:20px;font-weight:800;">{utr}</div>
                </div>
                <div style="text-align:center;">
                  <div style="color:#6b7280;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">Real Accepted</div>
                  <div style="color:#4ade80;font-size:20px;font-weight:800;">{acc}</div>
                </div>
                <div style="text-align:center;">
                  <div style="color:#6b7280;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">Sys. Deliveries</div>
                  <div style="color:#fbbf24;font-size:20px;font-weight:800;">{dl}</div>
                </div>
                <div style="text-align:center;">
                  <div style="color:#6b7280;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">Accept %</div>
                  <div style="color:#60a5fa;font-size:20px;font-weight:800;">{ar}%</div>
                </div>
                <div style="text-align:center;">
                  <div style="color:#6b7280;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">Balance</div>
                  <div style="color:#a78bfa;font-size:16px;font-weight:700;">{curr} {bal}</div>
                </div>
                <div>{bdg}</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

        # ── export button inside dashboard ────────────────────────────────────
        st.markdown("")
        xlsx = build_excel(riders)
        fname = f"riders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button("📥  Download Excel Report", data=xlsx, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — FULL TABLE
# ══════════════════════════════════════════════════════════════════════════════
with tab_table:
    riders = st.session_state.all_riders
    if not riders:
        st.markdown("""
        <div class="empty-state">
          <div class="empty-icon">📋</div>
          <div class="empty-title">No data yet</div>
          <div class="empty-sub">Extract riders first from the Upload tab.</div>
        </div>""", unsafe_allow_html=True)
    else:
        df = pd.DataFrame(riders)
        col_order = ["rider_id","name","phone","session","utr","deliveries","accepted",
                     "stacked","acceptance_rate","cash_balance","currency","rider_state","location_area"]
        df = df.reindex(columns=[c for c in col_order if c in df.columns])
        rename = {
            "rider_id":"ID","name":"Name","phone":"Phone","session":"Session",
            "utr":"UTR","deliveries":"Sys. Deliveries","accepted":"Real Accepted",
            "stacked":"Stacked","acceptance_rate":"Accept %","cash_balance":"Balance",
            "currency":"Currency","rider_state":"State","location_area":"Location"
        }
        df = df.rename(columns=rename)

        # filter bar
        fc1, fc2 = st.columns([3,1])
        with fc1:
            search = st.text_input("🔎 Search by name or ID", placeholder="Type to filter…", label_visibility="collapsed")
        with fc2:
            state_filter = st.selectbox("State", ["All","Working","Offline","Idle"], label_visibility="collapsed")

        if search:
            mask = df["Name"].str.contains(search, case=False, na=False) | \
                   df["ID"].astype(str).str.contains(search, na=False)
            df = df[mask]
        if state_filter != "All":
            df = df[df["State"].str.lower().str.contains(state_filter.lower(), na=False)]

        st.dataframe(df, use_container_width=True, height=500)

        st.markdown(f'<div style="color:#6b7280;font-size:12px;margin-top:8px;">Showing {len(df)} of {len(riders)} riders</div>', unsafe_allow_html=True)

st.markdown("</div>", unsafe_allow_html=True)
